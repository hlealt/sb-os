"""B3 Área do Investidor xlsx parser.

Parses the "Movimentação" export into orders, proventos, balcao,
and corporate_actions ledger rows.
"""

import re
from pathlib import Path

import openpyxl

from .base import InvestmentBaseParser, ParseResult, UnknownMapping, FlaggedOperation
from .name_map import NameMapResolver

# Ticker pattern: 4+ uppercase letters + digits, optionally with suffix
_TICKER_RE = re.compile(r'^([A-Z]{3,6}\d{1,2}[A-Z]?\d{0,2})\b')
# Option pattern: "Opção de Compra/Venda - TICKER - UNDERLYING"
_OPTION_RE = re.compile(r'^Opção de (?:Compra|Venda) - (\w+)')
# RF product pattern: "TYPE - CODE - ISSUER" or "TYPE - CODE"
_RF_RE = re.compile(r'^(CRA|CDB|DEB|LCA|LCI|LC|CRI)\s*-\s*(\S+)')
# Tesouro pattern
_TESOURO_RE = re.compile(r'^Tesouro\s+(.+)')

# Operations to ignore completely (paired Debito/Credito that net to zero,
# custody-only movements, or accounting-only entries)
_IGNORE_OPS = {
    'Atualização',
    'TRANSFERENCIA SEM FINANCEIRO',
    'TRANSFERÊNCIA SEM FINANCEIRO',
    'RETIRADA DE CUSTÓDIA',
    'Recibo de Subscrição',
    'Solicitação de Subscrição',
    'Direitos de Subscrição - Não Exercido',
    'Cessão de Direitos - Solicitada',
    'Direito de Subscrição',
    # Paired broker-to-broker custody transfers (same qty Debito+Credito same date)
    'Transferência',
    # JCP reassigned to a different broker — already captured on original broker
    'Juros Sobre Capital Próprio - Transferido',
}

# Operations to flag for manual review
_FLAG_OPS = {
    'Empréstimo/Aluguel de Ativos',
}

# RF asset types that route AMORTIZAÇÃO to balcao
_RF_TYPES = {'cra', 'deb', 'lca', 'lci', 'cdb', 'cdb_mp', 'tesouro', 'lc', 'cri', 'rf'}


class B3Parser(InvestmentBaseParser):
    source_id = "b3"

    def parse(self, filepath: Path, name_map: NameMapResolver = None) -> ParseResult:
        result = ParseResult()
        result.outputs = {
            'orders': [],
            'proventos': [],
            'balcao': [],
            'corporate': [],
        }

        rows = self._read_xlsx(filepath)

        for i, row in enumerate(rows):
            line_num = i + 2  # +1 for 0-index, +1 for header row
            es, date_str, mov, produto, instituicao, qty, price, valor = row

            # Skip empty rows
            if not mov:
                continue

            # Normalize Movimentação (strip trailing slashes, whitespace)
            mov_clean = mov.strip().rstrip('/')

            # Skip ignored operations
            if mov_clean in _IGNORE_OPS:
                continue

            # AMORTIZACAO PROGRAMADA: Debito is the accounting pair (skip);
            # Credito is the cash payment, treated identically to AMORTIZAÇÃO.
            if mov_clean == 'AMORTIZACAO PROGRAMADA':
                if es == 'Debito':
                    continue
                mov_clean = 'AMORTIZAÇÃO'

            # Skip PAGAMENTO DE JUROS Debito (accounting pair of the Credito)
            if mov_clean == 'PAGAMENTO DE JUROS' and es == 'Debito':
                continue

            # Flag operations
            if mov_clean in _FLAG_OPS:
                result.flags.append(FlaggedOperation(
                    date=self._parse_date(date_str),
                    operation=mov,
                    product=produto or '',
                    values={'qty': qty, 'price': price, 'valor': valor},
                    reason=f'Operation requires manual classification: {mov_clean}',
                    line_number=line_num,
                ))
                continue

            # Normalize product name
            canonical, asset_type = self._normalize_product(
                produto, name_map, result.unknowns, line_num
            )
            if canonical is None:
                continue  # Unknown product — added to unknowns list

            # Normalize institution → broker
            broker = self._normalize_institution(
                instituicao, name_map, result.unknowns, line_num
            )
            if broker is None:
                continue

            # Parse date
            date = self._parse_date(date_str)

            # Parse numeric values
            qty_val = self._parse_number(qty)
            price_val = self._parse_number(price)
            valor_val = self._parse_number(valor)

            # Route to appropriate ledger
            self._route_operation(
                result, mov_clean, es, date, canonical, asset_type,
                broker, qty_val, price_val, valor_val, line_num
            )

        return result

    def _read_xlsx(self, filepath: Path) -> list[tuple]:
        """Read B3 xlsx and return list of row tuples."""
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = []
        for i in range(2, ws.max_row + 1):  # Skip header
            row = tuple(ws.cell(row=i, column=j).value for j in range(1, 9))
            rows.append(row)
        wb.close()
        return rows

    def _parse_date(self, date_val) -> str:
        """Convert DD/MM/YYYY string or datetime to YYYY-MM-DD."""
        if date_val is None:
            return ''
        if hasattr(date_val, 'strftime'):
            return date_val.strftime('%Y-%m-%d')
        s = str(date_val).strip()
        parts = s.split('/')
        if len(parts) == 3:
            return f'{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}'
        return s

    def _parse_number(self, val) -> float:
        """Parse a numeric value, treating '-' and None as 0."""
        if val is None or val == '-' or val == '':
            return 0.0
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0

    def _normalize_product(self, produto: str, name_map: NameMapResolver,
                           unknowns: list, line_num: int) -> tuple[str | None, str]:
        """Extract canonical ticker/product_id and asset_type from B3 product string.

        Returns (canonical_value, asset_type) or (None, '') if unknown.
        """
        if not produto:
            return None, ''

        produto = produto.strip()

        # Check name_map first
        if name_map:
            result = name_map.resolve('b3', 'produto', produto)
            if result:
                return result['canonical_value'], result['asset_type']

        # Auto-extract for common patterns (fallback when name_map is empty)

        # Option: "Opção de Compra - EMBJE780 - EMBJ"
        m = _OPTION_RE.match(produto)
        if m:
            return m.group(1), 'opcao'

        # RF: "CRA - CODE - ISSUER" or "CDB - CODE"
        m = _RF_RE.match(produto)
        if m:
            rf_type = m.group(1).lower()
            code = m.group(2).rstrip(' -')
            # Build product_id: type_code (lowercase, underscores)
            product_id = f"{rf_type}_{code.lower().replace('-', '_')}"
            return product_id, rf_type

        # Tesouro: "Tesouro IPCA+ com Juros Semestrais 2032"
        m = _TESOURO_RE.match(produto)
        if m:
            name = m.group(1).strip()
            product_id = 'tesouro_' + re.sub(r'[^a-z0-9]+', '_', name.lower()).strip('_')
            return product_id, 'tesouro'

        # Listed asset: "TICKER - NAME"
        m = _TICKER_RE.match(produto)
        if m:
            ticker = m.group(1)
            # Determine type from ticker pattern
            if ticker.endswith('11') and len(ticker) >= 5:
                asset_type = 'fii'  # FIIs typically end in 11
            elif ticker.endswith(('39', '34', '32', '35')):
                asset_type = 'bdr'
            else:
                asset_type = 'acao'
            return ticker, asset_type

        # Unknown — add to unknowns list
        unknowns.append(UnknownMapping('b3', 'produto', produto, line_num))
        return None, ''

    def _normalize_institution(self, instituicao: str, name_map: NameMapResolver,
                               unknowns: list, line_num: int) -> str | None:
        """Normalize B3 institution name to canonical broker ID."""
        if not instituicao:
            return ''

        instituicao = instituicao.strip()

        # Check name_map
        if name_map:
            result = name_map.resolve_value('b3', 'instituicao', instituicao)
            if result:
                return result

        # Built-in mappings for common institutions
        inst_lower = instituicao.lower()
        if 'safra' in inst_lower:
            return 'safra'
        if 'clear' in inst_lower or 'nikos' in inst_lower:
            return 'clear'
        if 'xp' in inst_lower:
            return 'xp'
        if 'guide' in inst_lower:
            return 'safra'  # Guide migrated to Safra
        if 'avenue' in inst_lower:
            return 'avenue'
        if 'warren' in inst_lower:
            return 'warren'
        if 'mercado' in inst_lower:
            return 'mercado_pago'
        if 'bradesco' in inst_lower:
            return 'bradesco'

        unknowns.append(UnknownMapping('b3', 'instituicao', instituicao, line_num))
        return None

    def _route_operation(self, result: ParseResult, mov: str, es: str,
                         date: str, canonical: str, asset_type: str,
                         broker: str, qty: float, price: float, valor: float,
                         line_num: int):
        """Route a classified operation to the appropriate ledger."""

        # --- ORDERS ---
        # Transferência - Liquidação: settled stock/RF trade.
        # Credito=buy (share/position entered custody), Debito=sell.
        # Cessão de Direitos: subscription rights sold to another party → side V.
        if mov in ('Compra', 'Venda', 'Transferência - Liquidação', 'Cessão de Direitos'):
            if mov == 'Compra':
                side = 'C'
            elif mov == 'Venda':
                side = 'V'
            elif mov == 'Cessão de Direitos':
                side = 'V'
            else:
                side = 'C' if es == 'Credito' else 'V'
            if asset_type in _RF_TYPES:
                op = 'aplicacao' if side == 'C' else 'resgate'
                amt = -abs(valor) if side == 'C' else abs(valor)
                self._add_balcao(result, date, op, canonical, asset_type,
                                 broker, qty, amt if amt else 0)
            else:
                self._add_order(result, date, side, canonical, asset_type,
                                broker, qty, price, valor)
            return

        if mov in ('COMPRA/VENDA', 'COMPRA / VENDA',
                    'COMPRA/VENDA DEFINITIVA/CESSAO'):
            side = 'C' if es == 'Credito' else 'V'
            if asset_type in _RF_TYPES:
                op = 'aplicacao' if side == 'C' else 'resgate'
                amt = -abs(valor) if side == 'C' else abs(valor)
                self._add_balcao(result, date, op, canonical, asset_type,
                                 broker, qty, amt if amt else 0)
            else:
                self._add_order(result, date, side, canonical, asset_type,
                                broker, qty, price, valor)
            return

        if mov == 'Direitos de Subscrição - Exercido':
            self._add_order(result, date, 'C', canonical, asset_type,
                            broker, qty, price, valor)
            return

        if mov == 'Sobras de Subscrição':
            self._add_order(result, date, 'C', canonical, asset_type,
                            broker, qty, price, valor)
            return

        # Leilão de Fração: B3 auctions the fractional residue left over from a
        # corporate action and credits the proceeds as cash. Functionally
        # identical to "Fração em Ativos" — the listed qty was already rounded
        # to integer by the broker at the corporate event, so this row carries
        # NO qty delta, only cash. Routes to proventos type=fracao (NOT orders,
        # which would wrongly subtract the fraction from the rounded position).
        if mov == 'Leilão de Fração':
            self._add_provento(result, date, 'fracao', canonical,
                               broker, qty, price, valor)
            return

        # --- PROVENTOS ---
        if mov == 'Dividendo':
            self._add_provento(result, date, 'dividendo', canonical,
                               broker, qty, price, valor)
            return

        if mov == 'Juros Sobre Capital Próprio':
            self._add_provento(result, date, 'jcp', canonical,
                               broker, qty, price, valor)
            return

        if mov in ('Rendimento', 'Rendimento Líquido'):
            self._add_provento(result, date, 'rendimento', canonical,
                               broker, qty, price, valor)
            return

        # PAGAMENTO DE JUROS / Juros: RF income flows into balcao (so it nets
        # against aplicado for P&L), not proventos. Mirrors AMORTIZAÇÃO routing.
        if mov in ('PAGAMENTO DE JUROS', 'Juros'):
            if asset_type in _RF_TYPES:
                self._add_balcao(result, date, 'juros', canonical,
                                 asset_type, broker, qty, abs(valor))
            else:
                self._add_provento(result, date, 'juros', canonical,
                                   broker, qty, price, valor)
            return

        if mov == 'Bonificação em Dinheiro':
            self._add_provento(result, date, 'bonificacao_dinheiro', canonical,
                               broker, qty, price, valor)
            return

        # Fração em Ativos: cash payout for fractional shares after corporate actions
        if mov == 'Fração em Ativos':
            self._add_provento(result, date, 'fracao', canonical,
                               broker, qty, price, valor)
            return

        # --- AMORTIZAÇÃO: depends on asset_type ---
        if mov == 'AMORTIZAÇÃO':
            if asset_type in _RF_TYPES:
                self._add_balcao(result, date, 'amortizacao', canonical,
                                 asset_type, broker, qty, valor)
            else:
                self._add_provento(result, date, 'amortizacao', canonical,
                                   broker, qty, price, valor)
            return

        # --- RESTITUIÇÃO DE CAPITAL ---
        if mov == 'Restituição de Capital':
            if qty > 0 and price == 0 and valor == 0:
                # Shares returned → corporate action
                result.outputs['corporate'].append({
                    'date': date,
                    'action_type': 'restituicao_capital',
                    'ticker': canonical,
                    'new_ticker': '',
                    'ratio_from': '',
                    'ratio_to': '',
                    'broker': broker,
                    'source': 'b3',
                    'notes': f'qty={qty}',
                })
            else:
                # Cash → proventos
                self._add_provento(result, date, 'restituicao_capital', canonical,
                                   broker, qty, price, valor)
            return

        # --- BALCAO ---
        if mov == 'APLICAÇÃO':
            amount = -abs(valor) if valor else 0  # Application = money out (negative)
            self._add_balcao(result, date, 'aplicacao', canonical,
                             asset_type, broker, qty, amount)
            return

        if mov in ('RESGATE ANTECIPADO', 'RESGATE ANTECIPADO/'):
            amount = abs(valor) if valor else 0  # Redemption = money in (positive)
            self._add_balcao(result, date, 'resgate', canonical,
                             asset_type, broker, qty, amount)
            return

        if mov == 'VENCIMENTO':
            amount = abs(valor) if valor else 0
            self._add_balcao(result, date, 'vencimento', canonical,
                             asset_type, broker, qty, amount)
            return

        # --- CORPORATE ACTIONS ---
        if mov in ('Grupamento', 'Desdobramento', 'Cisão',
                    'Bonificação em Ativos', 'Incorporação',
                    'Conversão de Ativos', 'Fusão'):
            action_map = {
                'Grupamento': 'grupamento',
                'Desdobramento': 'desdobramento',
                'Cisão': 'cisao',
                'Bonificação em Ativos': 'bonificacao',
                'Incorporação': 'incorporacao',
                'Conversão de Ativos': 'conversao',
                'Fusão': 'fusao',
            }
            result.outputs['corporate'].append({
                'date': date,
                'action_type': action_map[mov],
                'ticker': canonical,
                'new_ticker': '',
                'ratio_from': '',
                'ratio_to': '',
                'broker': broker,
                'source': 'b3',
                'notes': f'qty={qty}',
            })
            return

        # --- UNCLASSIFIED ---
        result.flags.append(FlaggedOperation(
            date=date,
            operation=mov,
            product=canonical,
            values={'qty': qty, 'price': price, 'valor': valor},
            reason=f'Unclassified operation: {mov}',
            line_number=line_num,
        ))

    def _add_order(self, result: ParseResult, date: str, side: str,
                   ticker: str, asset_type: str, broker: str,
                   qty: float, price: float, valor: float):
        """Add a row to orders output."""
        # Calculate total (valor from B3 = qty × price, doesn't include fees)
        total = round(valor, 2) if valor else round(qty * price, 2)
        result.outputs['orders'].append({
            'date': date,
            'side': side,
            'ticker': ticker,
            'quantity': str(round(qty, 8)),
            'price': str(round(price, 8)),
            'currency': 'BRL',
            'total': str(total),
            'fees_exchange': '0',
            'fees_brokerage': '0',
            'fees_irrf': '0',
            'broker': broker,
            'asset_type': asset_type,
            'market': 'opcoes' if asset_type == 'opcao' else 'vista',
            'source': 'b3',
        })

    def _add_provento(self, result: ParseResult, date: str, ptype: str,
                      ticker: str, broker: str, qty: float, price: float,
                      valor: float):
        """Add a row to proventos output."""
        gross = round(valor, 2) if valor else round(qty * price, 2)
        vpu = round(price, 8) if price else (round(gross / qty, 8) if qty else 0)
        result.outputs['proventos'].append({
            'date': date,
            'type': ptype,
            'ticker': ticker,
            'quantity': str(round(qty, 8)),
            'value_per_unit': str(vpu),
            'gross_value': str(gross),
            'currency': 'BRL',
            'irrf': '0',
            'withholding_tax': '0',
            'net_value': str(gross),  # IRRF from B3 is 0; actual IRRF parsed from Safra PDFs
            'broker': broker,
            'source': 'b3',
        })

    def _add_balcao(self, result: ParseResult, date: str, operation: str,
                    canonical: str, asset_type: str, broker: str,
                    qty: float, amount: float):
        """Add a row to balcao output."""
        # CDB normalization: qty in R$0.01 units
        # Detect by checking if this looks like centavo-units (price ≈ 0.01)
        # For balcao, just store the monetary amount
        result.outputs['balcao'].append({
            'date': date,
            'operation': operation,
            'product_id': canonical,
            'product_type': asset_type if asset_type in _RF_TYPES else 'rf',
            'quantity': str(round(qty, 8)),
            'amount': str(round(amount, 2)),
            'irrf': '0',
            'iof': '0',
            'broker': broker,
            'source': 'b3',
        })
