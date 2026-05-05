"""Inspect B3 movimentação xlsx and list distinct RF/Tesouro 'produto' strings.

Output: each unique produto + the set of mov operations seen for it.
Used to drive name_map.csv entries (B3 raw produto → canonical balcão product_id).
"""

import openpyxl
from collections import defaultdict
from pathlib import Path


XLSX = Path('C:/Users/henri/Documents/second-brain/4-archives/finance/investments/historical-data/ok-b3-data/movimentacao-full.xlsx')

RF_PREFIXES = ('CRA', 'CDB', 'DEB', 'LCA', 'LCI', 'LC ', 'CRI', 'Tesouro')


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active

    products: dict[str, set] = defaultdict(set)

    for i in range(2, ws.max_row + 1):
        mov = ws.cell(row=i, column=3).value
        produto = ws.cell(row=i, column=4).value

        if not produto or not mov:
            continue

        produto = str(produto).strip()
        mov = str(mov).strip().rstrip('/')

        if not any(produto.startswith(p) for p in RF_PREFIXES):
            continue

        products[produto].add(mov)

    print(f'Distinct RF/Tesouro products: {len(products)}')
    print()
    for produto in sorted(products):
        movs = ' | '.join(sorted(products[produto]))
        print(f'{produto}')
        print(f'  movs: {movs}')
        print()


if __name__ == '__main__':
    main()
